import os
import zipfile
import io
import base64
import traceback
import logging

from PIL import Image
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from pathlib import Path
from typing import List, Tuple, Optional, Dict
import xml.etree.ElementTree as ET

from src.core.image_describer.base_image_describer import BaseImageDescriber
from src.core.llm_client import LLMClientFactory

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

XML_NAMESPACES = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
}


class ExcelImageDescriber(BaseImageDescriber):
    def __init__(self, llm_client=None):
        # Nếu không truyền llm_client thì tự khởi tạo Gemini
        if llm_client is None:
            factory = LLMClientFactory()
            llm_client = factory.get_client("gemini-2.5-flash")
        super().__init__(llm_client)
        self.client = llm_client["client"]
        self.model = llm_client["model"]

    def _load_workbook(self, file_path: str):
        """Load Excel workbook from file path."""
        try:
            logger.debug(f"Loading workbook from {file_path}")
            return load_workbook(file_path)
        except Exception as e:
            logger.error(f"Failed to load workbook: {str(e)}")
            raise ValueError(f"Failed to load Excel workbook: {str(e)}")

    def _build_drawing_relationships_map(self, zip_file: zipfile.ZipFile) -> Dict[str, Dict[str, str]]:
        """Build map of relationships between drawings and media."""
        drawing_rels_map = {}
        for rels in [f for f in zip_file.namelist() if f.startswith('xl/drawings/_rels/')]:
            root = ET.fromstring(zip_file.read(rels))
            rel_map = {}
            for rel in root:
                rid = rel.attrib['Id']
                target = rel.attrib['Target']
                if target.startswith("../media/"):
                    target = target.replace("../media/", "xl/media/")
                rel_map[rid] = target
            drawing_rels_map[os.path.basename(rels).replace('.rels', '')] = rel_map
        return drawing_rels_map

    def extract_images_and_contexts(self, file_path: str) -> List[Tuple[Image.Image, Optional[str], str, str]]:
        """Extract images from the xlsx (Linux-compatible)."""
        logger.info(f"Extracting images from {file_path}")
        image_info = []
        workbook = self._load_workbook(file_path)

        try:
            with zipfile.ZipFile(file_path, 'r') as zip_file:
                drawing_rels_map = self._build_drawing_relationships_map(zip_file)

                # map sheets to drawing.xml
                sheet_drawing_map = {}
                for srel in [f for f in zip_file.namelist() if f.startswith('xl/worksheets/_rels/')]:
                    sheet_name_xml = os.path.basename(srel).replace('.xml.rels', '')
                    root = ET.fromstring(zip_file.read(srel))
                    for r in root:
                        if 'drawing' in r.attrib.get('Type', ''):
                            drawing_target = os.path.basename(r.attrib['Target'])
                            sheet_drawing_map[sheet_name_xml] = drawing_target

                for idx, sheet in enumerate(workbook.worksheets, start=1):
                    sheet_key = f"sheet{idx}"
                    drawing_name = sheet_drawing_map.get(sheet_key)
                    if not drawing_name:
                        continue
                    drawing_xml = f"xl/drawings/{drawing_name}"
                    relmap = drawing_rels_map.get(drawing_name)
                    if not relmap or drawing_xml not in zip_file.namelist():
                        continue
                    root = ET.fromstring(zip_file.read(drawing_xml))
                    for anchor in root.findall('.//xdr:twoCellAnchor', XML_NAMESPACES):
                        row = int(anchor.find('xdr:from/xdr:row', XML_NAMESPACES).text) + 1
                        col = int(anchor.find('xdr:from/xdr:col', XML_NAMESPACES).text) + 1
                        cell = f"{get_column_letter(col)}{row}"
                        pic = anchor.find('.//xdr:pic', XML_NAMESPACES)
                        if pic is not None:
                            blip = pic.find('.//a:blip', XML_NAMESPACES)
                            if blip is not None:
                                rid = blip.attrib.get(f'{{{XML_NAMESPACES["r"]}}}embed')
                                target = relmap.get(rid)
                                if target and target in zip_file.namelist():
                                    img_data = zip_file.read(target)
                                    pil_img = Image.open(io.BytesIO(img_data)).convert("RGB")
                                    image_info.append((pil_img, None, sheet.title, cell))

            logger.info(f"Found {len(image_info)} images in workbook")
            return image_info
        except Exception as e:
            logger.error(f"Error extracting images: {str(e)}")
            raise

    def _convert_to_base64(self, image: Image.Image) -> str:
        """Convert PIL Image to base64 string."""
        buffered = io.BytesIO()
        image.save(buffered, format="PNG")
        img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')
        return img_str

    def describe_image(self, image: Image.Image, context: Optional[str] = None) -> str:
        """Sinh mô tả cho 1 ảnh bằng Gemini vision (chuẩn inline_data)."""
        try:
            base64_image = self._convert_to_base64(image)
            prompt = "Describe this image in detail about 20-30 words."
            if context:
                prompt += f"\nContext: {context}"
            logger.info(f"Generating description using model {self.model}")
            response = self.client.models.generate_content(
                model=self.model,
                contents=[
                    {
                        "role": "user",
                        "parts": [
                            {"text": prompt},
                            {"inline_data": {"mime_type": "image/png", "data": base64_image}}
                        ]
                    }
                ],
            )
            if hasattr(response, "candidates") and response.candidates:
                candidate = response.candidates[0]
                if candidate.content.parts:
                    # Log raw Gemini response for debugging
                    logger.info(f"Gemini raw response: {candidate.content.parts}")
                    description = candidate.content.parts[0].text.strip()
                    return description
                else:
                    logger.warning(f"Gemini candidate has no parts: {candidate}")
            else:
                logger.warning(f"Gemini response has no candidates: {response}")
            return "[No description generated]"
        except Exception as e:
            logger.error(f"Failed to generate description: {str(e)}")
            return f"[Failed to generate description: {str(e)}]"

    def generate_descriptions(self, image_context_pairs: List[Tuple[Image.Image, Optional[str]]]) -> List[str]:
        """Generate descriptions for images using LLM."""
        descriptions = []
        for img, ctx in image_context_pairs:
            descriptions.append(self.describe_image(img, ctx))
        return descriptions

    def replace_images_with_description(
        self,
        file_path: str,
        images_and_contexts: List[Tuple[Image.Image, Optional[str], str, str]],
        descriptions: List[str],
        output_path: str,
    ):
        """Insert description text into the Excel workbook near the image location."""
        logger.info(f"Replacing images with descriptions in {file_path}")
        try:
            wb = load_workbook(file_path)
            for (img, ctx, sheet_name, cell), desc in zip(images_and_contexts, descriptions):
                ws = wb[sheet_name]

                anchor_col = ''.join([c for c in cell if c.isalpha()])
                anchor_row = int(''.join([c for c in cell if c.isdigit()]))
                target_cell = f"{anchor_col}{anchor_row}"
                cell_obj = ws[target_cell]

                # Nếu là MergedCell → tìm top-left cell của merge range
                if isinstance(cell_obj, MergedCell):
                    for merged_range in ws.merged_cells.ranges:
                        if target_cell in merged_range:
                            top_left = merged_range.start_cell
                            logger.debug(f"Cell {target_cell} is merged. Writing to top-left {top_left.coordinate}")
                            cell_obj = top_left
                            break

                old_val = cell_obj.value or ""
                cell_obj.value = f"{old_val}\n[IMAGE] {desc}"

            logger.info(f"Saving workbook to {output_path}")
            wb.save(output_path)
        except Exception as e:
            logger.error(f"Error replacing images with descriptions: {str(e)}")
            raise

    def run(self, file_path: str, output_path: Optional[str] = None):
        """Process Excel file and add image descriptions."""
        try:
            logger.info(f"Processing {file_path}")

            if output_path is None:
                output_path = str(Path(file_path).with_name(Path(file_path).stem + "_with_desc.xlsx"))
            logger.debug(f"Output path: {output_path}")

            images_and_contexts = self.extract_images_and_contexts(file_path)
            if not images_and_contexts:
                logger.warning("No images found in workbook")
                return file_path

            logger.info("Generating descriptions for images")
            image_context_pairs = [(img, ctx) for img, ctx, _, _ in images_and_contexts]
            descriptions = self.generate_descriptions(image_context_pairs)

            logger.info("Replacing images with descriptions")
            self.replace_images_with_description(file_path, images_and_contexts, descriptions, output_path)

            logger.info(f"Successfully processed file. Output saved to: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error processing file: {str(e)}")
            logger.debug(traceback.format_exc())
            raise


if __name__ == "__main__":
    """Example usage of ExcelImageDescriber."""
    try:
        describer = ExcelImageDescriber()  # Luôn dùng Gemini
        input_file = "data/sample_documents/iiPay-Global-Payroll-Request-for-Proposal-Template-1.xlsx"
        output_file = "data/sample_documents/iiPay-Global-Payroll-Request-for-Proposal-Template-1_described.xlsx"
        result_path = describer.run(input_file, output_file)
        logger.info(f"Successfully processed file. Output saved to: {result_path}")
    except Exception as e:
        logger.error(f"Error in main: {str(e)}")
        logger.debug(traceback.format_exc())
